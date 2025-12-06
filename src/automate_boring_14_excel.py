# Notes made during chapter 14 of Automate the Boring Stuff With Python
# 2025-12-03

# %% Open an Excel workbook ---------------------------------------------------

import openpyxl
wb = openpyxl.load_workbook('file_name.xlsx')

type(wb)

# fetch specific sheet
wb.sheetnames
sheet = wb['Sheet3']
type(sheet)

sheet.title # = 'Sheet3'

# active sheet (last viewed)
another_sheet = wb.active

# fetch value
sheet['A1'] # a cell object
sheet['A1'].value # the value within the cell

# fetch position infromation
cell = sheet['B1']

f'Row {cell.row}, Column {cell.column} contains the value {cell.value}'
f'Cell {cell.coordinate} is {cell.value}'

# specify cell by row and column
sheet.cell(row = 1, column = 2)
# this is cell B2
# Excel is 1-indexed (like R) not 0-indexed (like Python)

# convert column numbers and letters
from openpyxl.utils import get_column_letter, column_index_from_string

get_column_letter(1) # 'A'
get_column_letter(27) # 'AA'

column_index_from_string('B') # 2
column_index_from_string('AB') # 28

# %% Fetch entire area --------------------------------------------------------

sheet_slice = sheet['A1':'C3']

for row_in_slice in sheet_slice:
    for cell in row_in_slice:
        print(cell.coordinate, cell.value)
    print('--- END OF ROW ---')

sheet.max_row
sheet.max_column

slice_start = 'A1'
slice_end = get_column_letter(sheet.max_column) + sheet.max_row

active_sheet_contents = sheet[slice_start:slice_end]

# %% Writing Excel files ------------------------------------------------------

# create empty workbook
wb = openpyxl.Workbook()
wb.sheetnames # workbook starts with one sheet
sheet = wb.active
sheet.title = 'new title' # change the title

# add and remove sheets
wb.create_sheet() # add new sheet
wb.create_sheet(index = 0, title = 'first sheet') # position sheet in first place

del wb['sheet_name'] # delete a sheet

# write values to cells
sheet['A1'] = 'new value'

# save
wb.save('file_name.xlsx')

# %% Advanced writing of Excel files ------------------------------------------

# specify font (size, bold, italic, name)
from openpyxl.styles import Font
font_settings = Font(size=12, italic=True)
sheet['A1'].font = font_settings

# formula as text string
sheet['B2'] = '=SUM(A1:A5)'
# but value will not be calculated until file is openned in Excel

# row height and column width
sheet.row_dimensions[1].height = 70
sheet.column_dimensions['B'].width = 50

# freeze panes
sheet.freeze_panes = 'A2'
# specify the first unfrozen cell
# So 'A2' = Row 1 frozen, no columns frozen

# charts are a six step process
# 1) select region of cells to plot
ref_obj = openpyxl.chart.Reference(sheet, 1, 1, 1, 10)
# 2) made region a series and name it
series_obj = openpyxl.chart.Series(ref_obj, title = "name")
# 3) make empty plot
chart_obj = openpyxl.chart.BarChart()
# 4) add title/labels/etc
chart_obj.title = 'Demo chart'
# 5) add series to chart
chart_obj.append(series_obj)
# 6) locate chart into the sheet
sheet.add_chart(chart_obj, 'C6')

# %%
